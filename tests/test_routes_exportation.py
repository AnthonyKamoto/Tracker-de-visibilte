"""
Tests unitaires pour les routes d'exportation (XLSX et CSV).
Tableau croise : une ligne par contenu avec colonnes appareils et navigateurs.
"""

import os
import sys
import io

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pytest
from openpyxl import load_workbook
from serveur.appli import creer_application


@pytest.fixture
def client(bdd_temporaire):
    """Client de test Flask avec des donnees de test."""
    appli = creer_application()
    appli.config['TESTING'] = True
    with appli.test_client() as c:
        c.post('/api/sessions', json={
            'id_session': 'export-s1',
            'type_appareil': 'ordinateur',
            'largeur_ecran': 1920,
            'hauteur_ecran': 1080,
            'navigateur': 'Chrome 120',
            'page_consultee': '/actualites'
        })
        c.post('/api/sessions', json={
            'id_session': 'export-s2',
            'type_appareil': 'mobile',
            'largeur_ecran': 375,
            'hauteur_ecran': 812,
            'navigateur': 'Safari 17',
            'page_consultee': '/actualites'
        })
        c.post('/api/evenements', json={
            'id_session': 'export-s1',
            'evenements': [
                {'id_contenu': 'banniere-1', 'type_contenu': 'banniere',
                 'pourcentage_visibilite': 0.8, 'duree_exposition_ms': 4000},
                {'id_contenu': 'banniere-2', 'type_contenu': 'banniere',
                 'pourcentage_visibilite': 0.5, 'duree_exposition_ms': 2000},
            ]
        })
        c.post('/api/evenements', json={
            'id_session': 'export-s2',
            'evenements': [
                {'id_contenu': 'banniere-1', 'type_contenu': 'banniere',
                 'pourcentage_visibilite': 1.0, 'duree_exposition_ms': 6000},
            ]
        })
        yield c


def test_export_xlsx(client):
    """GET /api/exportation/xlsx retourne un classeur avec un seul tableau croise."""
    reponse = client.get('/api/exportation/xlsx')
    assert reponse.status_code == 200
    assert 'spreadsheet' in reponse.content_type

    classeur = load_workbook(io.BytesIO(reponse.data))
    feuille = classeur['Statistiques']
    lignes = list(feuille.values)

    # En-tetes : colonnes fixes + appareils + navigateurs
    en_tetes = lignes[0]
    assert en_tetes[0] == 'ID Contenu'
    assert en_tetes[1] == 'Type'
    assert en_tetes[2] == 'Nombre de vues'
    assert 'mobile' in en_tetes
    assert 'ordinateur' in en_tetes
    assert 'Chrome 120' in en_tetes
    assert 'Safari 17' in en_tetes

    # 2 contenus (banniere-1 et banniere-2)
    assert len(lignes) == 3  # en-tete + 2 lignes de donnees


def test_export_xlsx_donnees_croisees(client):
    """Les colonnes appareils/navigateurs contiennent les bons comptages."""
    reponse = client.get('/api/exportation/xlsx')
    classeur = load_workbook(io.BytesIO(reponse.data))
    feuille = classeur['Statistiques']
    lignes = list(feuille.values)

    en_tetes = lignes[0]
    idx_ordinateur = en_tetes.index('ordinateur')
    idx_mobile = en_tetes.index('mobile')

    # banniere-1 : 1 vue ordinateur (export-s1) + 1 vue mobile (export-s2) = 2 vues
    # banniere-1 est en premiere ligne (plus de vues)
    ligne_b1 = lignes[1]
    assert ligne_b1[0] == 'banniere-1'
    assert ligne_b1[idx_ordinateur] == 1
    assert ligne_b1[idx_mobile] == 1

    # banniere-2 : 1 vue ordinateur (export-s1) + 0 mobile
    ligne_b2 = lignes[2]
    assert ligne_b2[0] == 'banniere-2'
    assert ligne_b2[idx_ordinateur] == 1
    assert ligne_b2[idx_mobile] == 0


def test_export_csv(client):
    """GET /api/exportation/csv retourne un seul fichier CSV avec tableau croise."""
    reponse = client.get('/api/exportation/csv')
    assert reponse.status_code == 200
    assert 'csv' in reponse.content_type

    contenu = reponse.data.decode('utf-8-sig')
    lignes = contenu.strip().split('\n')

    # En-tete + 2 lignes de donnees
    assert len(lignes) == 3

    assert 'ID Contenu' in lignes[0]
    assert 'ordinateur' in lignes[0]
    assert 'mobile' in lignes[0]
    assert 'Chrome 120' in lignes[0]
    assert 'banniere-1' in contenu


def test_export_xlsx_filtre_date_vide(client):
    """Export avec filtre excluant tout retourne un tableau vide (en-tetes seuls)."""
    reponse = client.get('/api/exportation/xlsx?date_debut=2099-01-01')
    assert reponse.status_code == 200

    classeur = load_workbook(io.BytesIO(reponse.data))
    feuille = classeur['Statistiques']
    lignes = list(feuille.values)
    # Seulement l'en-tete (5 colonnes fixes, pas d'appareils/navigateurs)
    assert len(lignes) == 1
    assert lignes[0][0] == 'ID Contenu'


def test_export_csv_filtre_date_vide(client):
    """Export CSV avec filtre excluant tout retourne seulement les en-tetes."""
    reponse = client.get('/api/exportation/csv?date_debut=2099-01-01')
    assert reponse.status_code == 200

    contenu = reponse.data.decode('utf-8-sig')
    lignes = contenu.strip().split('\n')
    assert len(lignes) == 1
    assert 'banniere' not in contenu


def test_export_xlsx_bdd_vide(bdd_temporaire):
    """Export XLSX avec BDD vide fonctionne sans erreur."""
    appli = creer_application()
    appli.config['TESTING'] = True
    with appli.test_client() as c:
        reponse = c.get('/api/exportation/xlsx')
        assert reponse.status_code == 200
        classeur = load_workbook(io.BytesIO(reponse.data))
        assert 'Statistiques' in classeur.sheetnames


def test_export_csv_bdd_vide(bdd_temporaire):
    """Export CSV avec BDD vide fonctionne sans erreur."""
    appli = creer_application()
    appli.config['TESTING'] = True
    with appli.test_client() as c:
        reponse = c.get('/api/exportation/csv')
        assert reponse.status_code == 200
        assert 'csv' in reponse.content_type
