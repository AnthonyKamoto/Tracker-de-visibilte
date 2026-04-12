"""
Routes API pour l'exportation des donnees en XLSX et CSV.
Tableau croise unique : une ligne par contenu, avec colonnes dynamiques
par type d'appareil et par navigateur.
"""

import os
import csv
import re
import logging

from flask import Blueprint, request, send_file, jsonify
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from serveur.config import REPERTOIRE_BASE
from serveur.base_de_donnees import obtenir_connexion

exportation_bp = Blueprint('exportation', __name__)
journal = logging.getLogger(__name__)

DOSSIER_EXPORTS = os.path.join(REPERTOIRE_BASE, 'exports')

_FORMAT_DATE = re.compile(r'^\d{4}-\d{2}-\d{2}$')


def _extraire_filtres_date():
    """Extrait les parametres date_debut et date_fin de la requete."""
    return request.args.get('date_debut'), request.args.get('date_fin')


def _assurer_dossier_exports():
    """Cree le dossier exports/ s'il n'existe pas."""
    os.makedirs(DOSSIER_EXPORTS, exist_ok=True)


def _clause_date(date_debut=None, date_fin=None):
    """Construit les conditions WHERE pour le filtrage par date."""
    conditions = []
    parametres = []
    if date_debut and _FORMAT_DATE.match(date_debut):
        conditions.append("e.date_enregistrement >= ?")
        parametres.append(date_debut)
    if date_fin and _FORMAT_DATE.match(date_fin):
        conditions.append("e.date_enregistrement <= ?")
        parametres.append(date_fin + " 23:59:59")
    return conditions, parametres


def _construire_tableau_croise(date_debut=None, date_fin=None):
    """Construit le tableau croise complet : contenus x appareils x navigateurs."""
    connexion = obtenir_connexion()
    try:
        curseur = connexion.cursor()
        conditions, parametres = _clause_date(date_debut, date_fin)
        clause_where = ""
        if conditions:
            clause_where = "WHERE " + " AND ".join(conditions)

        # Recuperer les types d'appareils existants
        curseur.execute(f"""
            SELECT DISTINCT s.type_appareil
            FROM evenements_visibilite e
            JOIN sessions s ON e.id_session = s.id_session
            {clause_where}
            ORDER BY s.type_appareil
        """, parametres)
        types_appareils = [ligne['type_appareil'] for ligne in curseur.fetchall()
                          if ligne['type_appareil']]

        # Recuperer les navigateurs existants
        curseur.execute(f"""
            SELECT DISTINCT s.navigateur
            FROM evenements_visibilite e
            JOIN sessions s ON e.id_session = s.id_session
            {clause_where}
            ORDER BY s.navigateur
        """, parametres)
        navigateurs = [ligne['navigateur'] for ligne in curseur.fetchall()
                      if ligne['navigateur']]

        # Construire la requete pivot avec CASE WHEN
        colonnes_appareils = ", ".join(
            f"SUM(CASE WHEN s.type_appareil = ? THEN 1 ELSE 0 END)"
            for _ in types_appareils
        )
        colonnes_navigateurs = ", ".join(
            f"SUM(CASE WHEN s.navigateur = ? THEN 1 ELSE 0 END)"
            for _ in navigateurs
        )

        colonnes_dynamiques = ""
        params_pivot = []
        if colonnes_appareils:
            colonnes_dynamiques += ", " + colonnes_appareils
            params_pivot.extend(types_appareils)
        if colonnes_navigateurs:
            colonnes_dynamiques += ", " + colonnes_navigateurs
            params_pivot.extend(navigateurs)

        requete = f"""
            SELECT
                e.id_contenu,
                e.type_contenu,
                COUNT(*) AS nombre_vues,
                ROUND(AVG(e.duree_exposition_ms), 0) AS duree_moyenne_ms,
                ROUND(AVG(e.pourcentage_visibilite), 2) AS visibilite_moyenne
                {colonnes_dynamiques}
            FROM evenements_visibilite e
            JOIN sessions s ON e.id_session = s.id_session
            {clause_where}
            GROUP BY e.id_contenu, e.type_contenu
            ORDER BY nombre_vues DESC
        """
        tous_parametres = params_pivot + parametres
        curseur.execute(requete, tous_parametres)

        nb_colonnes_fixes = 5
        lignes = []
        for row in curseur.fetchall():
            valeurs = list(row)
            # Convertir les comptages appareils/navigateurs en entiers
            for i in range(nb_colonnes_fixes, len(valeurs)):
                if valeurs[i] is not None:
                    valeurs[i] = int(valeurs[i])
            lignes.append(valeurs)

        # Construire les en-tetes
        en_tetes = ['ID Contenu', 'Type', 'Nombre de vues',
                    'Duree moyenne (ms)', 'Visibilite moyenne']
        en_tetes.extend(types_appareils)
        en_tetes.extend(navigateurs)

        return en_tetes, lignes

    finally:
        connexion.close()


@exportation_bp.route('/api/exportation/xlsx')
def exporter_xlsx():
    """Exporte le tableau croise dans un fichier Excel unique."""
    try:
        date_debut, date_fin = _extraire_filtres_date()
        en_tetes, lignes = _construire_tableau_croise(date_debut, date_fin)
        _assurer_dossier_exports()

        classeur = Workbook()
        feuille = classeur.active
        feuille.title = 'Statistiques'

        style_entete = Font(bold=True, color='FFFFFF')
        fond_entete = PatternFill(start_color='1A237E', end_color='1A237E',
                                  fill_type='solid')

        # Ecrire les en-tetes
        for i, en_tete in enumerate(en_tetes, 1):
            cellule = feuille.cell(row=1, column=i, value=en_tete)
            cellule.font = style_entete
            cellule.fill = fond_entete
            cellule.alignment = Alignment(horizontal='center')

        # Ecrire les donnees
        for num_ligne, ligne in enumerate(lignes, 2):
            for i, valeur in enumerate(ligne, 1):
                feuille.cell(row=num_ligne, column=i, value=valeur)

        # Ajuster la largeur des colonnes
        for colonne in feuille.columns:
            largeur_max = 0
            for cellule in colonne:
                if cellule.value is not None:
                    largeur_max = max(largeur_max, len(str(cellule.value)))
            lettre = colonne[0].column_letter
            feuille.column_dimensions[lettre].width = min(largeur_max + 4, 40)

        chemin_fichier = os.path.join(DOSSIER_EXPORTS, 'statistiques.xlsx')
        classeur.save(chemin_fichier)

        return send_file(
            chemin_fichier,
            as_attachment=True,
            download_name='statistiques.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as exc:
        journal.error("Erreur export XLSX : %s", exc)
        return jsonify({'succes': False, 'erreur': 'Erreur lors de l\'export XLSX'}), 500


@exportation_bp.route('/api/exportation/csv')
def exporter_csv():
    """Exporte le tableau croise dans un seul fichier CSV."""
    try:
        date_debut, date_fin = _extraire_filtres_date()
        en_tetes, lignes = _construire_tableau_croise(date_debut, date_fin)
        _assurer_dossier_exports()

        chemin_fichier = os.path.join(DOSSIER_EXPORTS, 'statistiques.csv')

        with open(chemin_fichier, 'w', newline='', encoding='utf-8-sig') as fichier:
            ecrivain = csv.writer(fichier)
            ecrivain.writerow(en_tetes)
            for ligne in lignes:
                ecrivain.writerow(ligne)

        return send_file(
            chemin_fichier,
            as_attachment=True,
            download_name='statistiques.csv',
            mimetype='text/csv; charset=utf-8'
        )
    except Exception as exc:
        journal.error("Erreur export CSV : %s", exc)
        return jsonify({'succes': False, 'erreur': 'Erreur lors de l\'export CSV'}), 500
